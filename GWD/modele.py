"""
=============================================================================
BALLET DES TRANSPORTEURS SUR LES QUAIS DE CHARGEMENT
Modélisation en Programme Linéaire en Nombres Mixtes (PLNM) avec PuLP
Cas Pédagogique — Green Wood Design (GWD) — ENIT 2026
=============================================================================

Données source : Données GWD 2026 VF.xlsx  —  Feuille : Données Transporteurs

Le fichier Excel peut être modifié librement (ajout / suppression de
transporteurs ou de camions) : le code se recalibre automatiquement.
  • N (nombre de camions) est lu depuis l'Excel
  • K (nombre à sélectionner) est déduit automatiquement : K = N - 3
  • Q (nombre de quais) = 2  ← modifier ici si besoin

Structure attendue de la feuille Excel :
  Bloc 1 — Transporteurs :
    | Transporteur | Score Éco (0-100) | Pénalité Retard (TND/min) | Camions |

  Bloc 2 — Camions (après une ligne d'entête "Camion") :
    | Camion | Heure d'Arrivée | Durée de chargement [min] | Date limite de départ |

Usage :
    pip install pulp openpyxl
    python ballet_transporteurs_pulp.py
=============================================================================
"""

import sys
import os

# =============================================================================
# CONFIGURATION  ← seul endroit à modifier si les règles métier changent
# =============================================================================
FICHIER_EXCEL = "Données GWD 2026 VF.xlsx"   # nom du fichier Excel (même dossier)
Q             = 2                              # nombre de quais
M             = 1440                           # Big-M = 24 h en minutes
# K est calculé automatiquement depuis N (voir plus bas)
# =============================================================================

TRANSPORTEURS: dict = {}
CAMIONS:       dict = {}
N = 0
K = 0


# ---------------------------------------------------------------------------
# 1. UTILITAIRES
# ---------------------------------------------------------------------------

def heure_to_min(valeur, base_h: int = 18) -> int:
    """
    Convertit une heure en minutes depuis base_h heures.
    Accepte : '20h00', '20H00', '20:00', fraction Excel (0.833…), int.
    Gestion du passage minuit : heures < base_h → lendemain.
    """
    if isinstance(valeur, float) and valeur < 1.0:
        # Format Excel time sérialisé (fraction de journée)
        total_min = int(round(valeur * 1440))
        h, m = divmod(total_min, 60)
    elif isinstance(valeur, (int, float)):
        val = int(valeur)
        if val > 100:
            return val          # déjà en minutes absolues
        h, m = val, 0
    else:
        s = str(valeur).strip().upper().replace("H", ":").replace(" ", "")
        parts = s.split(":")
        h = int(parts[0]) if parts[0].isdigit() else 0
        m = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 0

    minutes = h * 60 + m
    if h < base_h:
        minutes += 24 * 60      # lendemain
    return minutes - base_h * 60


def min_to_hhmm(m: float, base_h: int = 18) -> str:
    """Convertit des minutes depuis base_h en 'HHhMM'."""
    total = int(round(m)) + base_h * 60
    return f"{(total // 60) % 24:02d}h{total % 60:02d}"


# ---------------------------------------------------------------------------
# 2. CHARGEMENT DEPUIS EXCEL
# ---------------------------------------------------------------------------

def charger_depuis_excel() -> None:
    """
    Charge TRANSPORTEURS et CAMIONS depuis FICHIER_EXCEL.

    Détection automatique des deux blocs par leurs lignes d'entête :
      - Ligne dont col[0] == 'Transporteur'  → début bloc transporteurs
      - Ligne dont col[0] == 'Camion'        → début bloc camions
    Robuste aux lignes vides, espaces, variantes de casse.
    N et K sont recalculés à chaque appel.
    """
    global N, K, TRANSPORTEURS, CAMIONS

    try:
        import openpyxl
    except ImportError:
        print("[ERREUR] openpyxl non installé.  Exécutez : pip install openpyxl")
        sys.exit(1)

    # Recherche du fichier (dossier du script ou dossier courant)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    chemins_candidats = [
        os.path.join(script_dir, FICHIER_EXCEL),
        os.path.join(os.getcwd(),  FICHIER_EXCEL),
    ]
    chemin = next((p for p in chemins_candidats if os.path.isfile(p)), None)
    if chemin is None:
        print(f"[ERREUR] Fichier introuvable : '{FICHIER_EXCEL}'")
        print(f"         Cherché dans :")
        for p in chemins_candidats:
            print(f"           {p}")
        sys.exit(1)

    wb = openpyxl.load_workbook(chemin, data_only=True)

    # Recherche de la feuille (insensible à la casse ET aux accents)
    import unicodedata
    def normalise(s):
        """Supprime les accents et met en minuscules pour comparaison robuste."""
        return unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode().lower()

    feuille = next(
        (name for name in wb.sheetnames if "transporteur" in normalise(name)),
        None
    )
    if feuille is None:
        print(f"[ERREUR] Aucune feuille contenant 'Transporteur' trouvée dans '{FICHIER_EXCEL}'.")
        print(f"         Feuilles disponibles : {wb.sheetnames}")
        sys.exit(1)

    ws   = wb[feuille]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    TRANSPORTEURS.clear()
    CAMIONS.clear()
    camion_vers_trans: dict = {}

    # --- Détection des indices d'entête ---
    idx_trans  = None
    idx_camion = None
    for i, row in enumerate(rows):
        if row[0] is None:
            continue
        cell = str(row[0]).strip().lower()
        if cell == "transporteur" and idx_trans is None:
            idx_trans = i
        elif cell == "camion" and idx_camion is None:
            idx_camion = i

    if idx_trans is None:
        print("[ERREUR] Entête 'Transporteur' introuvable dans la feuille.")
        sys.exit(1)
    if idx_camion is None:
        print("[ERREUR] Entête 'Camion' introuvable dans la feuille.")
        sys.exit(1)

    # --- Lecture du bloc transporteurs ---
    for row in rows[idx_trans + 1 : idx_camion]:
        if row[0] is None:
            continue
        nom         = str(row[0]).strip()
        score       = row[1]
        penalite    = row[2]
        camions_str = str(row[3]).strip() if row[3] is not None else ""

        if not nom or not isinstance(score, (int, float)):
            continue

        TRANSPORTEURS[nom] = {
            "score":    int(score),
            "penalite": int(penalite),
        }
        for c in [c.strip() for c in camions_str.split(",") if c.strip()]:
            camion_vers_trans[c] = nom

    if not TRANSPORTEURS:
        print("[ERREUR] Aucun transporteur lu depuis l'Excel.")
        sys.exit(1)

    # --- Lecture du bloc camions ---
    for row in rows[idx_camion + 1 :]:
        if row[0] is None:
            continue
        cid = str(row[0]).strip()
        if not cid:
            continue

        arr, dur, lim = row[1], row[2], row[3]
        if arr is None or dur is None or lim is None:
            continue

        trans = camion_vers_trans.get(cid)
        if trans is None:
            print(f"[AVERTISSEMENT] Camion '{cid}' sans transporteur associé — ignoré.")
            continue

        CAMIONS[cid] = {
            "transporteur": trans,
            "S":  TRANSPORTEURS[trans]["score"],
            "P":  TRANSPORTEURS[trans]["penalite"],
            "Ta": heure_to_min(arr),
            "Td": int(dur),
            "Tl": heure_to_min(lim),
        }

    if not CAMIONS:
        print("[ERREUR] Aucun camion lu depuis l'Excel.")
        sys.exit(1)

    # --- Recalcul automatique de N et K ---
    N = len(CAMIONS)
    K = N - 3       # règle métier : sélectionner N-3 camions parmi N
                    # ← modifier cette ligne si la règle change

    print(f"[INFO] '{feuille}' chargée depuis '{os.path.basename(chemin)}'")
    print(f"       {len(TRANSPORTEURS)} transporteurs : {', '.join(TRANSPORTEURS)}")
    print(f"       {N} camions : {', '.join(CAMIONS)}")
    print(f"       K={K} camions à sélectionner, Q={Q} quais")


# ---------------------------------------------------------------------------
# 3. CONSTRUCTION DU MODÈLE PULP
# ---------------------------------------------------------------------------

def construire_modele():
    """Construit et retourne le modèle PuLP PLNM."""
    try:
        import pulp
    except ImportError:
        raise ImportError("PuLP non installé. Exécutez : pip install pulp")

    camions = list(CAMIONS.keys())
    quais   = list(range(1, Q + 1))

    prob = pulp.LpProblem("Ballet_Transporteurs_GWD", pulp.LpMaximize)

    # --- Variables ---
    u = pulp.LpVariable.dicts("u", camions, cat="Binary")
    x = pulp.LpVariable.dicts("x",
            [(i, j) for i in camions for j in quais], cat="Binary")
    d = pulp.LpVariable.dicts("d", camions, lowBound=0, cat="Continuous")
    r = pulp.LpVariable.dicts("r", camions, lowBound=0, cat="Continuous")

    paires = [
        (i, k, j)
        for idx_i, i in enumerate(camions)
        for k in camions[idx_i + 1:]
        for j in quais
    ]
    y = pulp.LpVariable.dicts("y", paires, cat="Binary")

    # --- Objectif : max Z = Σ Sᵢ·uᵢ − Σ Pᵢ·rᵢ − Σ (dᵢ − Taᵢ·uᵢ) ---
    prob += (
        pulp.lpSum(CAMIONS[i]["S"] * u[i] for i in camions)
        - pulp.lpSum(CAMIONS[i]["P"] * r[i] for i in camions)
        - pulp.lpSum(d[i] - CAMIONS[i]["Ta"] * u[i] for i in camions)
    ), "Z"

    # --- C1 : sélection exacte de K camions ---
    prob += pulp.lpSum(u[i] for i in camions) == K, "C1_selection_exacte"

    # --- C2 : affectation liée à la sélection ---
    for i in camions:
        prob += pulp.lpSum(x[i, j] for j in quais) == u[i], f"C2_{i}"

    # --- C3 : début de chargement après arrivée ---
    for i in camions:
        Ta_i = CAMIONS[i]["Ta"]
        prob += d[i] >= Ta_i * u[i], f"C3a_{i}"
        prob += d[i] <= M   * u[i],  f"C3b_{i}"

    # --- C4 : calcul du retard ---
    for i in camions:
        prob += r[i] >= d[i] + CAMIONS[i]["Td"] - CAMIONS[i]["Tl"], f"C4a_{i}"
        prob += r[i] <= M * u[i],                                    f"C4b_{i}"

    # --- C5 & C6 : non-chevauchement (Big-M) ---
    for (i, k, j) in paires:
        prob += (
            d[i] + CAMIONS[i]["Td"]
            <= d[k] + M*(1-y[i,k,j]) + M*(1-x[i,j]) + M*(1-x[k,j]),
            f"C5_{i}_{k}_q{j}"
        )
        prob += (
            d[k] + CAMIONS[k]["Td"]
            <= d[i] + M*y[i,k,j] + M*(1-x[i,j]) + M*(1-x[k,j]),
            f"C6_{k}_{i}_q{j}"
        )

    return prob, u, x, d, r, y, camions, quais


# ---------------------------------------------------------------------------
# 4. AFFICHAGE DES RÉSULTATS
# ---------------------------------------------------------------------------

def afficher_resultats(prob, u, x, d, r, y, camions, quais):
    import pulp

    statut = pulp.LpStatus[prob.status]
    print("\n" + "=" * 95)
    print("  RÉSULTATS — Ballet des Transporteurs GWD")
    print("=" * 95)
    print(f"  Statut solveur : {statut}")

    if prob.status != 1:
        print("  [ERREUR] Pas de solution optimale trouvée.")
        return

    Z = pulp.value(prob.objective)
    print(f"  Valeur objective Z* = {Z:.2f}")
    print()

    selectionnes = [i for i in camions if pulp.value(u[i]) > 0.5]
    exclus       = [i for i in camions if pulp.value(u[i]) < 0.5]
    print(f"  ✔  Sélectionnés ({len(selectionnes)}) : {', '.join(selectionnes)}")
    print(f"  ✘  Exclus       ({len(exclus)})  : {', '.join(exclus)}")

    # Largeurs de colonnes
    W = {
        "camion"  :  6, "trans"   : 13, "eco"     :  4,
        "arrivee" :  7, "debut"   :  7, "attente" :  7,
        "fin"     :  7, "limite"  :  7, "retard"  :  7, "penalite":  9,
    }

    def row_str(cam, trans, eco, arr, deb, att, fin, lim, ret, pen):
        return (
            f"  │  {str(cam):<{W['camion']}}  {str(trans):<{W['trans']}}  "
            f"{str(eco):>{W['eco']}}  {str(arr):>{W['arrivee']}}  "
            f"{str(deb):>{W['debut']}}  {str(att):>{W['attente']}}  "
            f"{str(fin):>{W['fin']}}  {str(lim):>{W['limite']}}  "
            f"{str(ret):>{W['retard']}}  {str(pen):>{W['penalite']}}"
        )

    header = row_str("Camion","Transporteur","Éco",
                     "Arrivée","Début","Attente","Fin","Limite","Retard","Pénalité")
    sep    = row_str(*["─"*w for w in W.values()])
    line_w = len(sep) - 4

    print()
    for j in quais:
        camions_quai = sorted(
            [i for i in selectionnes if pulp.value(x[i, j]) > 0.5],
            key=lambda i: pulp.value(d[i])
        )

        titre = f" QUAI {j} ({len(camions_quai)} camion{'s' if len(camions_quai) != 1 else ''}) "
        print(f"  ┌─{titre}{'─' * (line_w - len(titre) - 1)}")
        print(header)
        print(sep)

        for i in camions_quai:
            Ta_i    = CAMIONS[i]["Ta"]
            di      = pulp.value(d[i])
            ri      = pulp.value(r[i])
            attente = max(0.0, di - Ta_i)
            Pi      = CAMIONS[i]["P"]

            print(row_str(
                i,
                CAMIONS[i]["transporteur"],
                CAMIONS[i]["S"],
                min_to_hhmm(Ta_i),
                min_to_hhmm(di),
                f"{int(round(attente))}m",
                min_to_hhmm(di + CAMIONS[i]["Td"]),
                min_to_hhmm(CAMIONS[i]["Tl"]),
                f"{int(round(ri))}m",
                f"{int(round(Pi * ri))} ₮",
            ))

        print(f"  └{'─' * (line_w + 2)}")
        print()

    # Décomposition de Z*
    score_eco   = sum(CAMIONS[i]["S"] * pulp.value(u[i]) for i in camions)
    penalites   = sum(CAMIONS[i]["P"] * pulp.value(r[i]) for i in camions)
    attente_tot = sum(pulp.value(d[i]) - CAMIONS[i]["Ta"] * pulp.value(u[i]) for i in camions)

    print(f"  Décomposition de Z* :")
    print(f"    + Score éco total     : {score_eco:.0f}")
    print(f"    − Pénalités retard    : {penalites:.0f} TND")
    print(f"    − Temps attente total : {attente_tot:.0f} min")
    print(f"    {'═' * 34}")
    print(f"      Z*                  = {Z:.2f}")
    print()


# ---------------------------------------------------------------------------
# 5. POINT D'ENTRÉE
# ---------------------------------------------------------------------------

def main():
    # 1. Chargement des données depuis l'Excel
    charger_depuis_excel()
    print(f"\n[INFO] N={N}, K={K}, Q={Q}, M={M}\n")

    # 2. Construction du modèle
    print("[INFO] Construction du modèle PuLP...")
    try:
        prob, u, x, d, r, y, camions, quais = construire_modele()
    except ImportError as e:
        print(f"\n[ERREUR] {e}")
        n_bin = N + N*Q + N*(N-1)//2*Q
        print(f"\n  Structure du modèle :")
        print(f"    Variables binaires  : u({N}) + x({N}×{Q}) + y({N*(N-1)//2}×{Q}) = {n_bin}")
        print(f"    Variables continues : d({N}) + r({N}) = {2*N}")
        print(f"    Contraintes         : C1 à C8")
        sys.exit(1)

    import pulp
    print(f"[INFO] {len(prob.variables())} variables, {len(prob.constraints)} contraintes")

    # 3. Résolution
    print("[INFO] Résolution avec CBC...")
    prob.solve(pulp.PULP_CBC_CMD(msg=1, timeLimit=300))

    # 4. Affichage
    afficher_resultats(prob, u, x, d, r, y, camions, quais)


if __name__ == "__main__":
    main()