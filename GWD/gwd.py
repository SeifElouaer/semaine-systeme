"""
=============================================================================
APPLICATION STREAMLIT — Ballet des Transporteurs GWD
Fichier unique — pas de sous-dossiers requis
=============================================================================
Lancer avec :
    streamlit run app_ballet.py

Fichiers requis dans le MÊME dossier :
    app_ballet.py
    modele.py
    heuristique.py
    Données GWD 2026 VF.xlsx
=============================================================================
"""

import os, sys, importlib, tempfile, time
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px

# ---------------------------------------------------------------------------
# Import des modules solver (sans les modifier)
# ---------------------------------------------------------------------------
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
solver = importlib.import_module("modele")
heur   = importlib.import_module("heuristique")

# ---------------------------------------------------------------------------
# Config page
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="Ballet des Transporteurs — GWD",
    page_icon="🚛",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ---------------------------------------------------------------------------
# CSS global
# ---------------------------------------------------------------------------
st.markdown("""
<style>
    .main-title {
        font-size: 1.9rem; font-weight: 700; color: #1a5276;
        border-bottom: 3px solid #2e86c1; padding-bottom: 0.4rem;
        margin-bottom: 0.6rem;
    }
    .section-title {
        font-size: 1.05rem; font-weight: 600; color: #1a5276;
        margin-top: 1.2rem; margin-bottom: 0.4rem;
    }
    .kpi-box        { background:#eaf4fc; border-left:5px solid #2e86c1; border-radius:6px; padding:.7rem 1rem; margin-bottom:.4rem; }
    .kpi-box-green  { background:#eafaf1; border-left:5px solid #27ae60; border-radius:6px; padding:.7rem 1rem; margin-bottom:.4rem; }
    .kpi-box-orange { background:#fef9e7; border-left:5px solid #f39c12; border-radius:6px; padding:.7rem 1rem; margin-bottom:.4rem; }
    .kpi-box-purple { background:#f5eef8; border-left:5px solid #8e44ad; border-radius:6px; padding:.7rem 1rem; margin-bottom:.4rem; }
    .kpi-label { font-size:.74rem; color:#666; text-transform:uppercase; letter-spacing:.05em; }
    .kpi-value { font-size:1.5rem; font-weight:700; color:#1a5276; }
    .selected-badge { display:inline-block; background:#27ae60; color:white; border-radius:4px; padding:2px 8px; font-size:.8rem; margin:2px; }
    .excluded-badge { display:inline-block; background:#e74c3c; color:white; border-radius:4px; padding:2px 8px; font-size:.8rem; margin:2px; }
    /* ── Sidebar background ── */
    [data-testid="stSidebar"] { background:#1a2a3a !important; }

    /* ── All text in sidebar ── */
    [data-testid="stSidebar"] * { color:#e8f0fe !important; }

    /* ── Radio buttons label ── */
    [data-testid="stSidebar"] .stRadio label { color:#e8f0fe !important; }

    /* ── Radio option selected ── */
    [data-testid="stSidebar"] .stRadio [data-baseweb="radio"] [aria-checked="true"] + div p {
        color:#ffffff !important; font-weight:700;
    }

    /* ── Section titles inside sidebar ── */
    [data-testid="stSidebar"] h1,
    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3 { color:#90caf9 !important; }

    /* ── Caption / small text ── */
    [data-testid="stSidebar"] small,
    [data-testid="stSidebar"] .stCaption { color:#a0b4c8 !important; }

    /* ── Info box in sidebar ── */
    [data-testid="stSidebar"] .stAlert { background:#243447 !important; border-color:#2e86c1 !important; }
    [data-testid="stSidebar"] .stAlert p { color:#e8f0fe !important; }

    /* ── Number input labels ── */
    [data-testid="stSidebar"] .stNumberInput label { color:#cfe2f3 !important; }

    /* ── Slider labels & values ── */
    [data-testid="stSidebar"] .stSlider label { color:#cfe2f3 !important; }
    [data-testid="stSidebar"] .stSlider [data-testid="stTickBar"] { color:#90caf9 !important; }

    /* ── Checkbox label ── */
    [data-testid="stSidebar"] .stCheckbox label p { color:#cfe2f3 !important; }

    /* ── File uploader text ── */
    [data-testid="stSidebar"] .stFileUploader label { color:#cfe2f3 !important; }
    [data-testid="stSidebar"] .stFileUploader [data-testid="stFileUploaderDropzone"] {
        background:#243447 !important; border-color:#2e86c1 !important;
    }
    [data-testid="stSidebar"] .stFileUploader [data-testid="stFileUploaderDropzone"] * {
        color:#cfe2f3 !important;
    }

    /* ── Divider ── */
    [data-testid="stSidebar"] hr { border-color:#2e5272 !important; }
</style>
""", unsafe_allow_html=True)

COLORS = px.colors.qualitative.Set2

# =============================================================================
# UTILITAIRES COMMUNS
# =============================================================================

def min_to_hhmm(m, base_h=18):
    total = int(round(m)) + base_h * 60
    return f"{(total//60)%24:02d}h{total%60:02d}"

def kpi(col, label, value, suffix="", style=""):
    cls = f"kpi-box{'-'+style if style else ''}"
    col.markdown(
        f'<div class="{cls}"><div class="kpi-label">{label}</div>'
        f'<div class="kpi-value">{value}{suffix}</div></div>',
        unsafe_allow_html=True)

def badges_selection(sel, excl):
    st.markdown(
        "**Sélectionnés :** " + " ".join(f'<span class="selected-badge">{c}</span>' for c in sel),
        unsafe_allow_html=True)
    st.markdown(
        "**Exclus :** " + " ".join(f'<span class="excluded-badge">{c}</span>' for c in excl),
        unsafe_allow_html=True)

# =============================================================================
# CHARGEMENT EXCEL (commun aux deux pages)
# =============================================================================

@st.cache_data(show_spinner="Chargement Excel…")
def charger_donnees(file_bytes, _filename, fichier_defaut):
    import openpyxl, unicodedata
    def normalise(s):
        return unicodedata.normalize("NFD", s).encode("ascii","ignore").decode().lower()

    if file_bytes:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
    else:
        sd = os.path.dirname(os.path.abspath(__file__))
        candidates = [os.path.join(sd, fichier_defaut),
                      os.path.join(os.getcwd(), fichier_defaut)]
        tmp_path = next((p for p in candidates if os.path.isfile(p)), None)
        if not tmp_path:
            return None, None, f"Fichier '{fichier_defaut}' introuvable dans le dossier du script."

    try:
        wb      = openpyxl.load_workbook(tmp_path, data_only=True)
        feuille = next((n for n in wb.sheetnames if "transporteur" in normalise(n)), None)
        if not feuille:
            return None, None, f"Feuille 'Transporteurs' introuvable. Feuilles : {wb.sheetnames}"

        ws   = wb[feuille]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        transporteurs, camions, c2t = {}, {}, {}

        idx_t = idx_c = None
        for i, row in enumerate(rows):
            if not row[0]: continue
            cell = str(row[0]).strip().lower()
            if cell == "transporteur" and idx_t is None: idx_t = i
            elif cell == "camion"     and idx_c is None: idx_c = i

        for row in rows[idx_t+1:idx_c]:
            if not row[0]: continue
            nom, score, pen = str(row[0]).strip(), row[1], row[2]
            cstr = str(row[3]).strip() if row[3] else ""
            if not nom or not isinstance(score,(int,float)): continue
            transporteurs[nom] = {"score":int(score), "penalite":int(pen)}
            for c in [c.strip() for c in cstr.split(",") if c.strip()]:
                c2t[c] = nom

        for row in rows[idx_c+1:]:
            if not row[0]: continue
            cid = str(row[0]).strip()
            if not cid: continue
            arr, dur, lim = row[1], row[2], row[3]
            if None in (arr, dur, lim): continue
            trans = c2t.get(cid)
            if not trans: continue
            camions[cid] = {
                "transporteur": trans,
                "S": transporteurs[trans]["score"],
                "P": transporteurs[trans]["penalite"],
                "Ta": solver.heure_to_min(arr),
                "Td": int(dur),
                "Tl": solver.heure_to_min(lim),
            }
        return transporteurs, camions, None
    except Exception as e:
        return None, None, str(e)
    finally:
        if file_bytes and os.path.exists(tmp_path):
            os.unlink(tmp_path)

def apercu_donnees(trans, cams):
    with st.expander("📊 Aperçu des données Excel", expanded=False):
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**Transporteurs**")
            st.dataframe(pd.DataFrame([
                {"Transporteur":k, "Score Éco":v["score"], "Pénalité (TND/min)":v["penalite"]}
                for k,v in trans.items()]), use_container_width=True, hide_index=True)
        with c2:
            st.markdown("**Camions**")
            st.dataframe(pd.DataFrame([
                {"Camion":k, "Transporteur":v["transporteur"],
                 "Arrivée":min_to_hhmm(v["Ta"]), "Durée (min)":v["Td"],
                 "Limite":min_to_hhmm(v["Tl"])}
                for k,v in cams.items()]), use_container_width=True, hide_index=True)

# =============================================================================
# GRAPHIQUES COMMUNS
# =============================================================================

def gantt_chart(rows, quais, title="📅 Diagramme de Gantt"):
    trans_list = list({r["Transporteur"] for r in rows})
    cmap   = {t: COLORS[i%len(COLORS)] for i,t in enumerate(trans_list)}
    fig    = go.Figure()
    base_h = 18
    shown  = set()

    for row in rows:
        debut_abs = (row["d_min"] + base_h*60) % 1440
        quai_y    = f"Quai {row['Quai']}"
        color     = cmap[row["Transporteur"]]
        at        = row["Attente_min"]

        if at > 0:
            arr_abs = (row["Ta_min"] + base_h*60) % 1440
            fig.add_trace(go.Bar(
                x=[at], y=[quai_y], base=arr_abs, orientation="h",
                marker_color="rgba(180,180,180,0.35)", marker_line_width=0,
                showlegend=False, name="Attente",
                hovertemplate=f"<b>{row['Camion']}</b> — Attente: {int(at)}m<extra></extra>"))

        show_leg = row["Transporteur"] not in shown
        if show_leg: shown.add(row["Transporteur"])
        fig.add_trace(go.Bar(
            x=[row["Td_min"]], y=[quai_y], base=debut_abs, orientation="h",
            marker_color=color, marker_line_color="white", marker_line_width=1.5,
            name=row["Transporteur"], legendgroup=row["Transporteur"],
            showlegend=show_leg,
            hovertemplate=(f"<b>{row['Camion']}</b> ({row['Transporteur']})<br>"
                           f"Début : {row['Début']}<br>Fin : {row['Fin']}<br>"
                           f"Durée : {int(row['Td_min'])}m<br>Retard : {int(row['Retard_min'])}m<br>"
                           f"Pénalité : {int(row['Pénalité_TND'])} ₮<extra></extra>")))

        fig.add_annotation(
            x=debut_abs + row["Td_min"]/2, y=quai_y,
            text=f"<b>{row['Camion']}</b>", showarrow=False,
            font=dict(size=11, color="white"), xanchor="center", yanchor="middle")

    ticks = list(range(0, 1440, 60))
    fig.update_layout(
        barmode="overlay", title=title,
        xaxis=dict(tickvals=ticks, ticktext=[f"{h:02d}h00" for h in range(24)],
                   gridcolor="#eee",
                   range=[(min(r["d_min"] for r in rows)+base_h*60-30)%1440,
                          (max(r["Fin_min"] for r in rows)+base_h*60+60)%1440]),
        yaxis=dict(categoryorder="array",
                   categoryarray=[f"Quai {j}" for j in sorted(quais)]),
        height=280 + len(quais)*80, plot_bgcolor="white",
        legend=dict(title="Transporteur", orientation="h", y=-0.28, x=0.5, xanchor="center"),
        margin=dict(l=80, r=30, t=55, b=130))
    return fig

def bar_penalites_chart(rows, title="💰 Pénalités de retard (TND)"):
    df = pd.DataFrame(rows)
    df["lbl"] = df["Retard_min"].apply(lambda v: f"{int(v)}m")
    fig = px.bar(df, x="Camion", y="Pénalité_TND", color="Transporteur",
                 text="lbl", color_discrete_sequence=COLORS, title=title)
    fig.update_traces(textposition="outside")
    fig.update_layout(height=330, plot_bgcolor="white",
                      yaxis=dict(gridcolor="#eee"),
                      legend=dict(orientation="h", y=-0.28, x=0.5, xanchor="center"),
                      margin=dict(l=60, r=30, t=50, b=110))
    return fig

def bar_scores_chart(rows):
    df = pd.DataFrame(rows).sort_values("Éco", ascending=True)
    fig = px.bar(df, x="Éco", y="Camion", orientation="h", color="Transporteur",
                 text="Éco", color_discrete_sequence=COLORS,
                 title="🌿 Score écologique par camion sélectionné")
    fig.update_traces(textposition="outside")
    fig.update_layout(height=380, plot_bgcolor="white",
                      xaxis=dict(range=[0,110], gridcolor="#eee"),
                      legend=dict(orientation="h", y=-0.2, x=0.5, xanchor="center"),
                      margin=dict(l=60, r=30, t=50, b=80))
    return fig

def pie_attente_chart(rows):
    df = pd.DataFrame(rows)
    df_att = df[df["Attente_min"] > 0]
    if df_att.empty: return None
    fig = px.pie(df_att, values="Attente_min", names="Camion",
                 title="⏳ Répartition du temps d'attente",
                 color_discrete_sequence=COLORS, hole=0.4)
    fig.update_traces(textinfo="label+percent")
    fig.update_layout(height=350, margin=dict(l=30,r=30,t=50,b=30))
    return fig

def scatter_chart(rows):
    df = pd.DataFrame(rows)
    df["Arrivée_abs"] = df["Ta_min"].apply(lambda m: (m+18*60)%1440)
    df["Début_abs"]   = df["d_min"].apply(lambda m: (m+18*60)%1440)
    fig = px.scatter(df, x="Arrivée_abs", y="Début_abs",
                     color="Transporteur", size="Td_min", text="Camion",
                     color_discrete_sequence=COLORS,
                     title="🕐 Arrivée vs Début de chargement",
                     labels={"Arrivée_abs":"Heure d'arrivée","Début_abs":"Heure de début"})
    mn = min(df["Arrivée_abs"].min(), df["Début_abs"].min()) - 30
    mx = max(df["Arrivée_abs"].max(), df["Début_abs"].max()) + 30
    fig.add_shape(type="line", x0=mn,y0=mn,x1=mx,y1=mx,
                  line=dict(color="grey",dash="dash",width=1))
    ticks = list(range(0,1440,120))
    labels = [f"{h:02d}h00" for h in range(0,24,2)]
    fig.update_layout(height=360, plot_bgcolor="white",
                      xaxis=dict(tickvals=ticks, ticktext=labels, gridcolor="#eee"),
                      yaxis=dict(tickvals=ticks, ticktext=labels, gridcolor="#eee"),
                      legend=dict(orientation="h", y=-0.28, x=0.5, xanchor="center"),
                      margin=dict(l=60,r=30,t=50,b=110))
    fig.update_traces(textposition="top center", textfont_size=10)
    return fig

def tableau_quai(df, quais, key_prefix=""):
    tabs = st.tabs([f"Quai {j}" for j in quais])
    for tab, j in zip(tabs, quais):
        with tab:
            dq = df[df["Quai"]==j].sort_values("d_min")[[
                "Camion","Transporteur","Éco","Arrivée","Début",
                "Attente_min","Fin","Limite","Retard_min","Pénalité_TND"
            ]].rename(columns={"Attente_min":"Attente (m)",
                                "Retard_min":"Retard (m)",
                                "Pénalité_TND":"Pénalité (₮)"})
            dq["Attente (m)"]  = dq["Attente (m)"].apply(lambda v: f"{int(v)}m")
            dq["Retard (m)"]   = dq["Retard (m)"].apply(lambda v: f"{int(v)}m")
            dq["Pénalité (₮)"] = dq["Pénalité (₮)"].apply(lambda v: f"{int(v)} ₮")
            st.dataframe(dq, use_container_width=True, hide_index=True)

def decomposition_z(score_eco, penalites, att_tot, Z, label="Z*"):
    st.dataframe(pd.DataFrame({
        "Composante": ["+ Score éco total","− Pénalités retard",
                       "− Temps attente total", f"= {label}"],
        "Valeur": [f"{score_eco:.0f}", f"−{penalites:.0f} TND",
                   f"−{att_tot:.0f} min", f"{Z:.2f}"]}),
        use_container_width=False, hide_index=True)

# =============================================================================
# PAGE 1 — PLNM
# =============================================================================

@st.cache_data(show_spinner="Résolution PLNM (CBC)…")
def resoudre_plnm(_trans, _cam, K, Q, M=1440):
    try:
        import pulp
    except ImportError:
        return None, "PuLP non installé : pip install pulp"

    solver.TRANSPORTEURS.clear(); solver.TRANSPORTEURS.update(_trans)
    solver.CAMIONS.clear();       solver.CAMIONS.update(_cam)
    solver.N = len(_cam); solver.K = K; solver.Q = Q; solver.M = M

    try:
        prob, u, x, d, r, y, camions, quais = solver.construire_modele()
    except Exception as e:
        return None, str(e)

    prob.solve(pulp.PULP_CBC_CMD(msg=0, timeLimit=300))
    if prob.status != 1:
        return None, f"Pas de solution optimale ({pulp.LpStatus[prob.status]})"

    Z         = pulp.value(prob.objective)
    sel       = [i for i in camions if pulp.value(u[i]) > 0.5]
    excl      = [i for i in camions if pulp.value(u[i]) < 0.5]
    score_eco = sum(_cam[i]["S"] * pulp.value(u[i]) for i in camions)
    penalites = sum(_cam[i]["P"] * pulp.value(r[i]) for i in camions)
    att_tot   = sum(pulp.value(d[i]) - _cam[i]["Ta"]*pulp.value(u[i]) for i in camions)

    rows_res = []
    for i in sel:
        Ta = _cam[i]["Ta"]; di = pulp.value(d[i]); ri = pulp.value(r[i])
        quai_i = next(j for j in quais if pulp.value(x[i,j]) > 0.5)
        rows_res.append({
            "Camion":i, "Transporteur":_cam[i]["transporteur"],
            "Éco":_cam[i]["S"], "Quai":quai_i,
            "Ta_min":Ta, "d_min":di,
            "Fin_min":di+_cam[i]["Td"], "Tl_min":_cam[i]["Tl"],
            "Td_min":_cam[i]["Td"], "Attente_min":max(0.0,di-Ta),
            "Retard_min":ri, "Pénalité_TND":_cam[i]["P"]*ri,
            "Arrivée":min_to_hhmm(Ta), "Début":min_to_hhmm(di),
            "Fin":min_to_hhmm(di+_cam[i]["Td"]), "Limite":min_to_hhmm(_cam[i]["Tl"]),
        })

    return {"Z":Z,"sel":sel,"excl":excl,"score_eco":score_eco,
            "penalites":penalites,"att_tot":att_tot,"rows":rows_res,
            "n_vars":len(prob.variables()),"n_cons":len(prob.constraints)}, None


def page_plnm():
    st.markdown('<div class="main-title">🔢 Modèle PLNM — Optimisation Exacte (PuLP / CBC)</div>',
                unsafe_allow_html=True)
    st.caption("Programme Linéaire en Nombres Mixtes · Sélection · Affectation · Ordonnancement")

    with st.sidebar:
        st.markdown("### ⚙️ Paramètres PLNM")
        uploaded   = st.file_uploader("📂 Fichier Excel", type=["xlsx"], key="plnm_up")
        Q_val      = st.number_input("Quais (Q)", 1, 10, 2, key="plnm_Q")
        K_auto     = st.checkbox("K automatique (N−3)", value=True, key="plnm_Ka")
        file_bytes = uploaded.read() if uploaded else None

        trans, cams, err = charger_donnees(file_bytes, getattr(uploaded,"name",""),
                                           solver.FICHIER_EXCEL)
        N_val = len(cams) if cams else 12
        if K_auto:
            K_val = max(1, N_val-3)
            if cams: st.info(f"K = {N_val} − 3 = **{K_val}**")
        else:
            K_val = st.number_input("K", 1, N_val, max(1,N_val-3), key="plnm_K")

        run_btn = st.button("▶ Lancer PLNM", use_container_width=True,
                            type="primary", key="plnm_run")

    if err:   st.error(f"❌ {err}"); return
    if not trans:
        st.info(f"👈 Placez '{solver.FICHIER_EXCEL}' dans le dossier ou uploadez-le."); return

    apercu_donnees(trans, cams)

    if run_btn or "plnm_res" in st.session_state:
        if run_btn:
            res, err2 = resoudre_plnm(trans, cams, K_val, Q_val)
            if err2: st.error(f"❌ {err2}"); return
            st.session_state["plnm_res"] = res
        res   = st.session_state["plnm_res"]
        rows  = res["rows"]
        quais = sorted({r["Quai"] for r in rows})
        df    = pd.DataFrame(rows)

        # KPIs
        st.markdown("---")
        st.markdown('<div class="section-title">📈 Indicateurs clés</div>', unsafe_allow_html=True)
        c1,c2,c3,c4,c5 = st.columns(5)
        kpi(c1,"Valeur Z*",           f"{res['Z']:.0f}")
        kpi(c2,"Score éco total",     f"{res['score_eco']:.0f}",  style="green")
        kpi(c3,"Pénalités retard",    f"{res['penalites']:.0f}", " TND", style="orange")
        kpi(c4,"Temps attente total", f"{res['att_tot']:.0f}",   " min")
        kpi(c5,"Camions retenus",     f"{len(res['sel'])}/{len(cams)}")

        st.markdown('<div class="section-title">✅ Sélection des camions</div>', unsafe_allow_html=True)
        badges_selection(res["sel"], res["excl"])

        st.markdown("---")
        st.plotly_chart(gantt_chart(rows, quais), use_container_width=True)

        st.markdown('<div class="section-title">📋 Détail par quai</div>', unsafe_allow_html=True)
        tableau_quai(df, quais, "plnm")

        st.markdown("---")
        st.markdown('<div class="section-title">📊 Analyses</div>', unsafe_allow_html=True)
        ca, cb = st.columns(2)
        with ca: st.plotly_chart(bar_scores_chart(rows), use_container_width=True)
        with cb: st.plotly_chart(bar_penalites_chart(rows), use_container_width=True)
        cc, cd = st.columns(2)
        with cc:
            p = pie_attente_chart(rows)
            st.plotly_chart(p, use_container_width=True) if p else st.info("✅ Aucune attente.")
        with cd: st.plotly_chart(scatter_chart(rows), use_container_width=True)

        st.markdown("---")
        st.markdown('<div class="section-title">🧮 Décomposition de Z*</div>', unsafe_allow_html=True)
        decomposition_z(res["score_eco"], res["penalites"], res["att_tot"], res["Z"])
        st.caption(f"PLNM · {res['n_vars']} variables · {res['n_cons']} contraintes · CBC")
    else:
        st.info("👈 Configurez les paramètres puis cliquez sur **▶ Lancer PLNM**.")


# =============================================================================
# PAGE 2 — HEURISTIQUES
# =============================================================================

def run_heuristiques(trans, cams, K, Q,
                     taille_pop, n_gen, taux_crois, taux_mut,
                     elitisme, patience, graine):
    heur.TRANSPORTEURS.clear(); heur.TRANSPORTEURS.update(trans)
    heur.CAMIONS.clear();       heur.CAMIONS.update(cams)
    heur.N = len(cams); heur.K = K; heur.Q = Q

    t0 = time.time()
    sol_greedy          = heur.greedy()
    Z_greedy, _         = heur.evaluer(sol_greedy)
    sol_ag, Z_ag, hist  = heur.algorithme_genetique(
        solution_initiale=sol_greedy,
        taille_pop=taille_pop, n_generations=n_gen,
        taux_croisement=taux_crois, taux_mutation=taux_mut,
        taille_tournoi=3, elitisme=elitisme,
        patience=patience, graine=graine if graine >= 0 else None)
    duree = time.time() - t0

    def extract(solution, Z):
        sel    = solution["selection"]
        aff    = solution["quais"]
        ordres = solution.get("_ordres", {})
        rows   = []
        for q in range(1, Q+1):
            ordre_q = ordres.get(q, [c for c in sel if aff[c]==q])
            tl = 0
            for c in ordre_q:
                debut = max(heur.CAMIONS[c]["Ta"], tl)
                fin   = debut + heur.CAMIONS[c]["Td"]
                ri    = max(0.0, fin - heur.CAMIONS[c]["Tl"])
                at    = debut - heur.CAMIONS[c]["Ta"]
                Pi    = heur.CAMIONS[c]["P"]
                rows.append({
                    "Camion":c, "Transporteur":heur.CAMIONS[c]["transporteur"],
                    "Éco":heur.CAMIONS[c]["S"], "Quai":q,
                    "Ta_min":heur.CAMIONS[c]["Ta"], "d_min":debut,
                    "Fin_min":fin, "Tl_min":heur.CAMIONS[c]["Tl"],
                    "Td_min":heur.CAMIONS[c]["Td"],
                    "Attente_min":at, "Retard_min":ri, "Pénalité_TND":Pi*ri,
                    "Arrivée":min_to_hhmm(heur.CAMIONS[c]["Ta"]),
                    "Début":min_to_hhmm(debut),
                    "Fin":min_to_hhmm(fin),
                    "Limite":min_to_hhmm(heur.CAMIONS[c]["Tl"]),
                })
                tl = fin
        excl      = [c for c in heur.CAMIONS if c not in sel]
        score_eco = sum(heur.CAMIONS[c]["S"] for c in sel)
        penalites = sum(r["Pénalité_TND"] for r in rows)
        att_tot   = sum(r["Attente_min"]   for r in rows)
        return {"Z":Z,"sel":sorted(sel),"excl":sorted(excl),"rows":rows,
                "score_eco":score_eco,"penalites":penalites,"att_tot":att_tot}

    rg = extract(sol_greedy, Z_greedy)
    ra = extract(sol_ag,     Z_ag)
    amelioration = Z_ag - Z_greedy
    pct = (amelioration/abs(Z_greedy)*100) if Z_greedy != 0 else 0
    return {"greedy":rg,"ag":ra,"historique":hist,
            "amelioration":amelioration,"pct":pct,"duree":duree}

def convergence_chart(historique):
    df = pd.DataFrame({"Génération":list(range(len(historique))), "Z":historique})
    fig = px.line(df, x="Génération", y="Z",
                  title="📈 Convergence — Algorithme Génétique",
                  color_discrete_sequence=["#8e44ad"])
    fig.add_hline(y=historique[-1], line_dash="dot", line_color="#27ae60",
                  annotation_text=f"Z final = {historique[-1]:.1f}",
                  annotation_position="bottom right")
    fig.update_layout(height=340, plot_bgcolor="white",
                      xaxis=dict(gridcolor="#eee"),
                      yaxis=dict(gridcolor="#eee"),
                      margin=dict(l=60,r=30,t=50,b=50))
    fig.update_traces(line_width=2)
    return fig

def comparaison_bar(Z_greedy, Z_ag):
    fig = go.Figure(go.Bar(
        x=["🔵 Greedy","🟣 Algorithme Génétique"],
        y=[Z_greedy, Z_ag],
        marker_color=["#2980b9","#8e44ad"],
        text=[f"{Z_greedy:.1f}", f"{Z_ag:.1f}"],
        textposition="outside"))
    fig.update_layout(title="⚖️ Comparaison Z : Greedy vs AG",
                      height=320, plot_bgcolor="white",
                      yaxis=dict(gridcolor="#eee"),
                      margin=dict(l=60,r=30,t=50,b=50))
    return fig

def page_heuristique():
    st.markdown('<div class="main-title">🧬 Heuristiques — Greedy + Algorithme Génétique</div>',
                unsafe_allow_html=True)
    st.caption("Phase 1 : construction gloutonne · Phase 2 : évolution génétique · Optimisation ordre quais")

    with st.sidebar:
        st.markdown("### ⚙️ Paramètres Heuristiques")
        uploaded   = st.file_uploader("📂 Fichier Excel", type=["xlsx"], key="heur_up")
        Q_val      = st.number_input("Quais (Q)", 1, 10, 2, key="heur_Q")
        K_auto     = st.checkbox("K automatique (N−3)", value=True, key="heur_Ka")
        file_bytes = uploaded.read() if uploaded else None

        trans, cams, err = charger_donnees(file_bytes, getattr(uploaded,"name",""),
                                           heur.FICHIER_EXCEL)
        N_val = len(cams) if cams else 12
        if K_auto:
            K_val = max(1, N_val-3)
            if cams: st.info(f"K = {N_val} − 3 = **{K_val}**")
        else:
            K_val = st.number_input("K", 1, N_val, max(1,N_val-3), key="heur_K")

        st.markdown("**Paramètres AG**")
        taille_pop = st.slider("Population",          20,  200,  80,  10, key="heur_pop")
        n_gen      = st.slider("Générations max",     50,  500, 300,  50, key="heur_gen")
        taux_crois = st.slider("Taux croisement",    0.5,  1.0, 0.85, 0.05, key="heur_crois")
        taux_mut   = st.slider("Taux mutation",      0.1,  0.9, 0.30, 0.05, key="heur_mut")
        elitisme   = st.slider("Élitisme",             1,   10,   3,   1,  key="heur_elit")
        patience   = st.slider("Patience",            20,  200,  80,  10, key="heur_pat")
        graine     = st.number_input("Graine (−1=aléatoire)", -1, 9999, 42, key="heur_seed")

        run_btn = st.button("▶ Lancer Heuristiques", use_container_width=True,
                            type="primary", key="heur_run")

    if err:   st.error(f"❌ {err}"); return
    if not trans:
        st.info(f"👈 Placez '{heur.FICHIER_EXCEL}' dans le dossier ou uploadez-le."); return

    apercu_donnees(trans, cams)

    if run_btn:
        with st.spinner("⏳ Greedy + Algorithme Génétique en cours…"):
            result = run_heuristiques(
                trans, cams, K_val, Q_val,
                taille_pop, n_gen, taux_crois, taux_mut,
                elitisme, patience, int(graine))
        st.session_state["heur_res"] = result
        st.success(f"✅ Terminé en {result['duree']:.2f}s")

    if "heur_res" not in st.session_state:
        st.info("👈 Configurez les paramètres puis cliquez sur **▶ Lancer Heuristiques**.")
        return

    result = st.session_state["heur_res"]
    rg     = result["greedy"]
    ra     = result["ag"]
    hist   = result["historique"]
    quais  = sorted({r["Quai"] for r in ra["rows"]})

    # --- Résumé comparatif ---
    st.markdown("---")
    st.markdown('<div class="section-title">⚖️ Résumé comparatif</div>', unsafe_allow_html=True)
    c1,c2,c3,c4 = st.columns(4)
    kpi(c1,"Z Greedy",               f"{rg['Z']:.1f}")
    kpi(c2,"Z Algorithme Génétique", f"{ra['Z']:.1f}", style="purple")
    kpi(c3,"Amélioration (Δ)",       f"{result['amelioration']:+.1f}", style="green")
    kpi(c4,"Temps total",            f"{result['duree']:.2f}", " s")

    ca, cb = st.columns(2)
    with ca: st.plotly_chart(comparaison_bar(rg["Z"], ra["Z"]), use_container_width=True)
    with cb: st.plotly_chart(convergence_chart(hist), use_container_width=True)

    # --- Onglets par phase ---
    st.markdown("---")
    tab_g, tab_a = st.tabs(["🔵 Phase 1 — Greedy", "🟣 Phase 2 — Algorithme Génétique"])

    for tab, res, label in [(tab_g, rg, "Greedy"), (tab_a, ra, "Algorithme Génétique")]:
        with tab:
            rows  = res["rows"]
            df    = pd.DataFrame(rows)
            style = "purple" if label != "Greedy" else ""

            c1,c2,c3,c4 = st.columns(4)
            kpi(c1,"Z",                   f"{res['Z']:.2f}",         style=style)
            kpi(c2,"Score éco",           f"{res['score_eco']:.0f}", style="green")
            kpi(c3,"Pénalités retard",    f"{res['penalites']:.0f}", " TND", style="orange")
            kpi(c4,"Temps attente total", f"{res['att_tot']:.0f}",   " min")

            st.markdown('<div class="section-title">✅ Sélection</div>', unsafe_allow_html=True)
            badges_selection(res["sel"], res["excl"])

            st.plotly_chart(gantt_chart(rows, quais, f"📅 Gantt — {label}"),
                            use_container_width=True)

            st.markdown('<div class="section-title">📋 Détail par quai</div>', unsafe_allow_html=True)
            tableau_quai(df, quais, label)

            ca, cb = st.columns(2)
            with ca: st.plotly_chart(bar_penalites_chart(rows, f"💰 Pénalités — {label}"),
                                     use_container_width=True)
            with cb:
                p = pie_attente_chart(rows)
                st.plotly_chart(p, use_container_width=True) if p else st.info("✅ Aucune attente.")

            st.markdown('<div class="section-title">🧮 Décomposition de Z</div>', unsafe_allow_html=True)
            decomposition_z(res["score_eco"], res["penalites"], res["att_tot"],
                            res["Z"], f"Z ({label})")


# =============================================================================
# NAVIGATION PRINCIPALE
# =============================================================================

def main():
    with st.sidebar:
        st.image("https://img.icons8.com/fluency/96/truck.png", width=58)
        st.markdown("## 🚛 GWD — Transporteurs")
        st.markdown("---")
        page = st.radio(
            "Navigation",
            options=["🔢 Modèle PLNM (Exact)", "🧬 Heuristiques (Greedy + AG)"],
            label_visibility="collapsed",
            key="nav_page",
        )
        st.markdown("---")
        st.caption("Green Wood Design · ENIT 2026")

    if page == "🔢 Modèle PLNM (Exact)":
        page_plnm()
    else:
        page_heuristique()


if __name__ == "__main__":
    main()