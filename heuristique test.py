"""
=============================================================================
BALLET DES TRANSPORTEURS SUR LES QUAIS DE CHARGEMENT
Méthodes Heuristiques : Greedy + Algorithme Génétique
Cas Pédagogique — Green Wood Design (GWD) — ENIT 2026
=============================================================================

Données source : Données GWD 2026 VF.xlsx  —  Feuille : Données Transporteurs

Le fichier Excel peut être modifié librement (ajout / suppression de
transporteurs ou de camions) : le code se recalibre automatiquement.
  • N (nombre de camions) est lu depuis l'Excel
  • K (nombre à sélectionner) est déduit automatiquement : K = N - 3
  • Q (nombre de quais) = 2  ← modifier ici si besoin

Structure attendue de la feuille Excel :
  Bloc 1 — Transporteurs :
    | Transporteur | Score Éco (0-100) | Pénalité Retard (TND/min) | Camions |

  Bloc 2 — Camions (après une ligne d'entête "Camion") :
    | Camion | Heure d'Arrivée | Durée de chargement [min] | Date limite de départ |

DEUX PHASES HEURISTIQUES :
  Phase 1 — Greedy (Glouton) :
    Construit rapidement une solution initiale de qualité en sélectionnant
    à chaque étape le meilleur (camion, quai) selon un score composite.

  Phase 2 — Algorithme Génétique (AG) :
    Améliore itérativement la solution par évolution d'une population.
    Opérateurs : sélection tournoi, croisement OX, mutation swap/quai.
    La fonction d'évaluation optimise aussi l'ORDRE de passage sur chaque
    quai (règles EDD, SPT, FIFO, Slack + exhaustif pour ≤ 4 camions).

OBJECTIF (identique au modèle PL) :
  Z = Σ Sᵢ·uᵢ  −  Σ Pᵢ·rᵢ  −  Σ (dᵢ − Taᵢ)
      ↑ score éco   ↑ pénalités   ↑ attente

Usage :
    pip install openpyxl
    python ballet_heuristique_gwd.py
=============================================================================
"""

import sys
import os
import random
import time
import copy
from typing import List, Tuple, Dict, Optional
from itertools import permutations as iter_perms

# =============================================================================
# CONFIGURATION  ← seul endroit à modifier si les règles métier changent
# =============================================================================
FICHIER_EXCEL = "Données GWD 2026 VF.xlsx"   # nom du fichier Excel (même dossier)
Q             = 2                              # nombre de quais
# K est calculé automatiquement depuis N (voir charger_depuis_excel)
# =============================================================================

TRANSPORTEURS: dict = {}
CAMIONS:       dict = {}
N = 0
K = 0


# =============================================================================
# 1. UTILITAIRES
# =============================================================================

def heure_to_min(valeur, base_h: int = 18) -> int:
    """
    Convertit une heure en minutes depuis base_h heures.
    Accepte : '20h00', '20H00', '20:00', fraction Excel (0.833…), int.
    Gestion du passage minuit : heures < base_h → lendemain.
    """
    if isinstance(valeur, float) and valeur < 1.0:
        # Format Excel time sérialisé (fraction de journée)
        total_min = int(round(valeur * 1440))
        h, m = divmod(total_min, 60)
    elif isinstance(valeur, (int, float)):
        val = int(valeur)
        if val > 100:
            return val          # déjà en minutes absolues
        h, m = val, 0
    else:
        s = str(valeur).strip().upper().replace("H", ":").replace(" ", "")
        parts = s.split(":")
        h = int(parts[0]) if parts[0].isdigit() else 0
        m = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 0

    minutes = h * 60 + m
    if h < base_h:
        minutes += 24 * 60      # lendemain
    return minutes - base_h * 60


def min_to_hhmm(m: float, base_h: int = 18) -> str:
    """Convertit des minutes depuis base_h en 'HHhMM'."""
    total = int(round(m)) + base_h * 60
    return f"{(total // 60) % 24:02d}h{total % 60:02d}"


# =============================================================================
# 2. CHARGEMENT DEPUIS EXCEL
# =============================================================================

def charger_depuis_excel() -> None:
    """
    Charge TRANSPORTEURS et CAMIONS depuis FICHIER_EXCEL.

    Détection automatique des deux blocs par leurs lignes d'entête :
      - Ligne dont col[0] == 'Transporteur'  → début bloc transporteurs
      - Ligne dont col[0] == 'Camion'        → début bloc camions
    Robuste aux lignes vides, espaces, variantes de casse.
    N et K sont recalculés à chaque appel.
    """
    global N, K, TRANSPORTEURS, CAMIONS

    try:
        import openpyxl
    except ImportError:
        print("[ERREUR] openpyxl non installé.  Exécutez : pip install openpyxl")
        sys.exit(1)

    # Recherche du fichier (dossier du script ou dossier courant)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    chemins_candidats = [
        os.path.join(script_dir, FICHIER_EXCEL),
        os.path.join(os.getcwd(),  FICHIER_EXCEL),
    ]
    chemin = next((p for p in chemins_candidats if os.path.isfile(p)), None)
    if chemin is None:
        print(f"[ERREUR] Fichier introuvable : '{FICHIER_EXCEL}'")
        print(f"         Cherché dans :")
        for p in chemins_candidats:
            print(f"           {p}")
        sys.exit(1)

    wb = openpyxl.load_workbook(chemin, data_only=True)

    # Recherche de la feuille (insensible à la casse ET aux accents)
    import unicodedata
    def normalise(s):
        return unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode().lower()

    feuille = next(
        (name for name in wb.sheetnames if "transporteur" in normalise(name)),
        None
    )
    if feuille is None:
        print(f"[ERREUR] Aucune feuille contenant 'Transporteur' trouvée dans '{FICHIER_EXCEL}'.")
        print(f"         Feuilles disponibles : {wb.sheetnames}")
        sys.exit(1)

    ws   = wb[feuille]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    TRANSPORTEURS.clear()
    CAMIONS.clear()
    camion_vers_trans: dict = {}

    # --- Détection des indices d'entête ---
    idx_trans  = None
    idx_camion = None
    for i, row in enumerate(rows):
        if row[0] is None:
            continue
        cell = str(row[0]).strip().lower()
        if cell == "transporteur" and idx_trans is None:
            idx_trans = i
        elif cell == "camion" and idx_camion is None:
            idx_camion = i

    if idx_trans is None:
        print("[ERREUR] Entête 'Transporteur' introuvable dans la feuille.")
        sys.exit(1)
    if idx_camion is None:
        print("[ERREUR] Entête 'Camion' introuvable dans la feuille.")
        sys.exit(1)

    # --- Lecture du bloc transporteurs ---
    for row in rows[idx_trans + 1 : idx_camion]:
        if row[0] is None:
            continue
        nom         = str(row[0]).strip()
        score       = row[1]
        penalite    = row[2]
        camions_str = str(row[3]).strip() if row[3] is not None else ""

        if not nom or not isinstance(score, (int, float)):
            continue

        TRANSPORTEURS[nom] = {
            "score":    int(score),
            "penalite": int(penalite),
        }
        for c in [c.strip() for c in camions_str.split(",") if c.strip()]:
            camion_vers_trans[c] = nom

    if not TRANSPORTEURS:
        print("[ERREUR] Aucun transporteur lu depuis l'Excel.")
        sys.exit(1)

    # --- Lecture du bloc camions ---
    for row in rows[idx_camion + 1 :]:
        if row[0] is None:
            continue
        cid = str(row[0]).strip()
        if not cid:
            continue

        arr, dur, lim = row[1], row[2], row[3]
        if arr is None or dur is None or lim is None:
            continue

        trans = camion_vers_trans.get(cid)
        if trans is None:
            print(f"[AVERTISSEMENT] Camion '{cid}' sans transporteur associé — ignoré.")
            continue

        CAMIONS[cid] = {
            "transporteur": trans,
            "S":  TRANSPORTEURS[trans]["score"],
            "P":  TRANSPORTEURS[trans]["penalite"],
            "Ta": heure_to_min(arr),
            "Td": int(dur),
            "Tl": heure_to_min(lim),
        }

    if not CAMIONS:
        print("[ERREUR] Aucun camion lu depuis l'Excel.")
        sys.exit(1)

    # --- Recalcul automatique de N et K ---
    N = len(CAMIONS)
    K = N - 3       # règle métier : sélectionner N-3 camions parmi N
                    # ← modifier cette ligne si la règle change

    print(f"[INFO] '{feuille}' chargée depuis '{os.path.basename(chemin)}'")
    print(f"       {len(TRANSPORTEURS)} transporteurs : {', '.join(TRANSPORTEURS)}")
    print(f"       {N} camions : {', '.join(CAMIONS)}")
    print(f"       K={K} camions à sélectionner, Q={Q} quais")


# =============================================================================
# 3. ÉVALUATION — AVEC OPTIMISATION DE L'ORDRE SUR LES QUAIS
# =============================================================================
#
#  Pour une affectation donnée (camion → quai), au lieu de simplement trier
#  par heure d'arrivée, on teste plusieurs règles de séquencement et on
#  garde le meilleur ordre pour chaque quai :
#
#  Règles testées :
#   1. EDD   (Earliest Due Date)    → trier par Tl croissant     ★ réduit les retards
#   2. SPT   (Shortest Processing)  → trier par Td croissant       libère vite le quai
#   3. FIFO  (First In First Out)   → trier par Ta croissant       intuitif
#   4. Slack (Tl - Td)              → trier par marge croissante   équilibre
#   5. Priorité pénalité            → trier par P décroissant      coût d'abord
#   6. Exhaustif (si ≤ 4 camions)   → toutes les permutations      optimal garanti

def meilleur_ordre(camions_quai: List[str]) -> Tuple[List[str], Dict]:
    """
    Retourne l'ordre de passage optimal parmi plusieurs règles de séquencement.
    Pour les petits groupes (≤ 4 camions), l'énumération exhaustive garantit
    l'optimalité locale.
    """
    if not camions_quai:
        return [], {}

    candidats = [
        sorted(camions_quai, key=lambda c: CAMIONS[c]["Tl"]),                     # EDD
        sorted(camions_quai, key=lambda c: CAMIONS[c]["Td"]),                     # SPT
        sorted(camions_quai, key=lambda c: CAMIONS[c]["Ta"]),                     # FIFO
        sorted(camions_quai, key=lambda c: CAMIONS[c]["Tl"] - CAMIONS[c]["Td"]), # Slack
        sorted(camions_quai, key=lambda c: CAMIONS[c]["P"], reverse=True),        # Pénalité
    ]

    # Exhaustif pour les petits groupes
    if len(camions_quai) <= 4:
        for perm in iter_perms(camions_quai):
            candidats.append(list(perm))

    meilleur_Z_local  = float("-inf")
    meilleur_ordre_local = candidats[0]
    meilleurs_details    = {}

    for ordre in candidats:
        details, _ = _simuler_quai(ordre, 0)
        se = sum(CAMIONS[c]["S"]              for c in ordre)
        pe = sum(CAMIONS[c]["P"] * details[c]["r"]  for c in ordre)
        at = sum(details[c]["attente"]        for c in ordre)
        Z_local = se - pe - at
        if Z_local > meilleur_Z_local:
            meilleur_Z_local     = Z_local
            meilleur_ordre_local = ordre
            meilleurs_details    = details

    return meilleur_ordre_local, meilleurs_details


def _simuler_quai(ordre: List[str], temps_libre: float) -> Tuple[Dict, float]:
    """Simule le passage d'une liste ordonnée de camions sur un quai."""
    details = {}
    tl = temps_libre
    for c in ordre:
        debut = max(CAMIONS[c]["Ta"], tl)
        fin   = debut + CAMIONS[c]["Td"]
        details[c] = {
            "d":      debut,
            "r":      max(0.0, fin - CAMIONS[c]["Tl"]),
            "attente": debut - CAMIONS[c]["Ta"],
            "fin":    fin,
        }
        tl = fin
    return details, tl


def evaluer(solution: Dict) -> Tuple[float, Dict]:
    """
    Calcule la valeur objective Z et les détails de planification.

    Pour chaque quai, l'ordre de passage est optimisé via meilleur_ordre().
    L'ordre retenu est sauvegardé dans solution['_ordres'] pour l'affichage.
    """
    sel         = solution["selection"]
    affectation = solution["quais"]

    details_global  = {}
    ordres_optimaux = {}

    for q in range(1, Q + 1):
        camions_quai = [c for c in sel if affectation[c] == q]
        if not camions_quai:
            ordres_optimaux[q] = []
            continue

        ordre_opt, _ = meilleur_ordre(camions_quai)
        ordres_optimaux[q] = ordre_opt

        # Recalcul avec continuité du quai (temps_libre cumulatif)
        tl = 0
        for c in ordre_opt:
            debut = max(CAMIONS[c]["Ta"], tl)
            fin   = debut + CAMIONS[c]["Td"]
            details_global[c] = {
                "d":      debut,
                "r":      max(0.0, fin - CAMIONS[c]["Tl"]),
                "attente": debut - CAMIONS[c]["Ta"],
                "fin":    fin,
            }
            tl = fin

    solution["_ordres"] = ordres_optimaux

    score_eco   = sum(CAMIONS[c]["S"]                      for c in sel)
    penalites   = sum(CAMIONS[c]["P"] * details_global[c]["r"] for c in sel)
    attente_tot = sum(details_global[c]["attente"]         for c in sel)
    Z = score_eco - penalites - attente_tot

    return Z, details_global


# =============================================================================
# 4. PHASE 1 — HEURISTIQUE GREEDY (GLOUTON)
# =============================================================================
#
# PRINCIPE :
#   On construit la solution en K étapes. À chaque étape, on choisit la
#   combinaison (camion, quai) qui maximise le score composite :
#     score = S_i  −  P_i × retard_estimé  −  attente_estimée  +  0.1 × slack
#   Le slack (marge = Tl - fin estimée) favorise les camions encore récupérables.
#
# COMPLEXITÉ : O(K × N × Q) — quasi-instantané.

def greedy() -> Dict:
    """Construit une solution initiale par heuristique gloutonne."""
    print("\n" + "─" * 65)
    print("  PHASE 1 — GREEDY (Construction de la solution initiale)")
    print("─" * 65)

    disponibles = list(CAMIONS.keys())
    selection   = []
    affectation = {}
    quai_libre  = {q: 0 for q in range(1, Q + 1)}  # quand chaque quai est disponible

    for etape in range(K):
        meilleur_camion = None
        meilleur_quai   = None
        meilleur_score  = float("-inf")

        for c in disponibles:
            Ta = CAMIONS[c]["Ta"]
            Td = CAMIONS[c]["Td"]
            Tl = CAMIONS[c]["Tl"]
            S  = CAMIONS[c]["S"]
            P  = CAMIONS[c]["P"]

            for q in range(1, Q + 1):
                debut  = max(Ta, quai_libre[q])
                fin    = debut + Td
                retard = max(0.0, fin - Tl)
                attente = debut - Ta
                slack  = max(0, Tl - fin)
                score  = S - P * retard - attente + 0.1 * slack

                if score > meilleur_score:
                    meilleur_score  = score
                    meilleur_camion = c
                    meilleur_quai   = q

        selection.append(meilleur_camion)
        affectation[meilleur_camion] = meilleur_quai

        Ta = CAMIONS[meilleur_camion]["Ta"]
        Td = CAMIONS[meilleur_camion]["Td"]
        debut = max(Ta, quai_libre[meilleur_quai])
        quai_libre[meilleur_quai] = debut + Td

        disponibles.remove(meilleur_camion)
        print(f"  Étape {etape+1:2d} : {meilleur_camion} → Quai {meilleur_quai}"
              f"  (score = {meilleur_score:.1f})")

    solution = {"selection": selection, "quais": affectation}
    Z, _     = evaluer(solution)
    print(f"\n  ✔ Solution greedy  →  Z = {Z:.2f}")
    return solution


# =============================================================================
# 5. PHASE 2 — ALGORITHME GÉNÉTIQUE
# =============================================================================
#
# ENCODAGE :
#   Un individu = {"selection": [K camions], "quais": {camion: quai}}
#
# OPÉRATEURS :
#   - Sélection par tournoi  : tirer k concurrents, garder le meilleur.
#   - Croisement OX          : préserve l'ordre relatif des gènes parents.
#   - Mutation swap          : échange un camion sélectionné ↔ exclu.
#   - Mutation quai          : change le quai d'un camion sélectionné.
#   - Élitisme               : les meilleurs individus sont conservés tels quels.

def _generer_individu_aleatoire() -> Dict:
    """Génère une solution aléatoire valide."""
    sel = random.sample(list(CAMIONS.keys()), K)
    aff = {c: random.randint(1, Q) for c in sel}
    return {"selection": sel, "quais": aff}


def _selection_tournoi(population: List[Dict], scores: List[float],
                       taille_tournoi: int = 3) -> Dict:
    """
    Sélection par tournoi.
    On tire aléatoirement taille_tournoi individus et on retourne le meilleur.
    """
    idx = random.sample(range(len(population)), taille_tournoi)
    return copy.deepcopy(population[max(idx, key=lambda i: scores[i])])


def _croisement_ox(parent1: Dict, parent2: Dict) -> Tuple[Dict, Dict]:
    """
    Croisement OX (Order Crossover).

    On copie une fenêtre [a, b] du parent source dans l'enfant,
    puis on complète dans l'ordre du second parent en évitant les doublons.
    Les quais sont hérités du parent dont provient chaque camion.
    """
    sel1, sel2 = parent1["selection"], parent2["selection"]
    aff1, aff2 = parent1["quais"],     parent2["quais"]

    a = random.randint(0, K - 2)
    b = random.randint(a + 1, K - 1)

    def _faire_enfant(s_src, s_fill, a_src, a_fill):
        enfant_sel = [None] * K
        fenetre    = set(s_src[a:b+1])
        enfant_sel[a:b+1] = s_src[a:b+1]
        pos = (b + 1) % K
        for i in range(K):
            c = s_fill[(b + 1 + i) % K]
            if c not in fenetre:
                enfant_sel[pos] = c
                fenetre.add(c)
                pos = (pos + 1) % K
        enfant_aff = {}
        for i, c in enumerate(enfant_sel):
            src = a_src if a <= i <= b else a_fill
            enfant_aff[c] = src.get(c, random.randint(1, Q))
        return {"selection": enfant_sel, "quais": enfant_aff}

    return _faire_enfant(sel1, sel2, aff1, aff2), \
           _faire_enfant(sel2, sel1, aff2, aff1)


def _muter(individu: Dict, taux_mutation: float = 0.3) -> Dict:
    """
    Mutation avec probabilité taux_mutation.
    Deux types au choix aléatoire :
      - Swap  : remplace un camion sélectionné par un non-sélectionné.
      - Quai  : change l'affectation d'un camion au quai.
    """
    individu = copy.deepcopy(individu)
    if random.random() > taux_mutation:
        return individu

    sel = individu["selection"]
    aff = individu["quais"]

    if random.random() < 0.5:
        non_sel = [c for c in CAMIONS if c not in sel]
        if non_sel:
            idx_out = random.randint(0, K - 1)
            c_out   = sel[idx_out]
            c_in    = random.choice(non_sel)
            del aff[c_out]
            sel[idx_out] = c_in
            aff[c_in]    = random.randint(1, Q)
    else:
        c = random.choice(sel)
        aff[c] = random.randint(1, Q)

    return individu


def algorithme_genetique(
    solution_initiale : Dict,
    taille_pop        : int   = 80,
    n_generations     : int   = 300,
    taux_croisement   : float = 0.85,
    taux_mutation     : float = 0.30,
    taille_tournoi    : int   = 3,
    elitisme          : int   = 3,
    patience          : int   = 80,
    graine            : Optional[int] = 42,
) -> Tuple[Dict, float, List[float]]:
    """
    Algorithme Génétique principal.

    Paramètres :
      taille_pop      : individus dans la population
      n_generations   : nombre maximum de générations
      taux_croisement : probabilité de croisement entre deux parents
      taux_mutation   : probabilité de muter un enfant
      taille_tournoi  : concurrents lors de la sélection
      elitisme        : meilleurs individus conservés tels quels à chaque génération
      patience        : arrêt anticipé si pas d'amélioration depuis N générations
      graine          : graine aléatoire (reproductibilité)
    """
    if graine is not None:
        random.seed(graine)

    print("\n" + "─" * 65)
    print("  PHASE 2 — ALGORITHME GÉNÉTIQUE (Amélioration)")
    print("─" * 65)
    print(f"  Population : {taille_pop}  |  Générations max : {n_generations}"
          f"  |  Croisement : {taux_croisement*100:.0f}%"
          f"  |  Mutation : {taux_mutation*100:.0f}%")

    # --- Initialisation de la population ---
    # La solution greedy sert de graine : 30 % des individus sont des mutations
    # légères de celle-ci, 70 % sont aléatoires pour assurer la diversité.
    population = [solution_initiale]
    for _ in range(taille_pop - 1):
        if random.random() < 0.3:
            population.append(_muter(solution_initiale, taux_mutation=1.0))
        else:
            population.append(_generer_individu_aleatoire())

    scores = [evaluer(ind)[0] for ind in population]

    meilleur_Z   = max(scores)
    meilleure_sol = copy.deepcopy(population[scores.index(meilleur_Z)])
    historique    = [meilleur_Z]
    stagnation    = 0

    t_debut = time.time()

    for gen in range(n_generations):
        indices_tries = sorted(range(taille_pop), key=lambda i: scores[i], reverse=True)

        nouvelle_pop = []

        # Élitisme : copier les `elitisme` meilleurs sans modification
        for i in range(elitisme):
            nouvelle_pop.append(copy.deepcopy(population[indices_tries[i]]))

        # Génération du reste de la population
        while len(nouvelle_pop) < taille_pop:
            p1 = _selection_tournoi(population, scores, taille_tournoi)
            p2 = _selection_tournoi(population, scores, taille_tournoi)

            if random.random() < taux_croisement:
                e1, e2 = _croisement_ox(p1, p2)
            else:
                e1, e2 = copy.deepcopy(p1), copy.deepcopy(p2)

            e1 = _muter(e1, taux_mutation)
            e2 = _muter(e2, taux_mutation)
            nouvelle_pop.append(e1)
            if len(nouvelle_pop) < taille_pop:
                nouvelle_pop.append(e2)

        population = nouvelle_pop
        scores     = [evaluer(ind)[0] for ind in population]

        gen_max = max(scores)
        if gen_max > meilleur_Z:
            meilleur_Z    = gen_max
            meilleure_sol = copy.deepcopy(population[scores.index(gen_max)])
            stagnation    = 0
        else:
            stagnation += 1

        historique.append(meilleur_Z)

        if (gen + 1) % 50 == 0 or gen == 0:
            duree = time.time() - t_debut
            moy   = sum(scores) / len(scores)
            print(f"  Gén. {gen+1:>4d}/{n_generations}"
                  f"  |  Z_max = {meilleur_Z:>9.2f}"
                  f"  |  Z_moy = {moy:>9.2f}"
                  f"  |  {duree:.2f}s")

        # Arrêt anticipé si stagnation
        if stagnation >= patience:
            print(f"  [Convergence] Arrêt à la génération {gen+1} (stagnation {patience} gén.).")
            break

    duree_totale = time.time() - t_debut
    print(f"\n  ✔ AG terminé en {duree_totale:.3f}s  →  Z = {meilleur_Z:.2f}")
    return meilleure_sol, meilleur_Z, historique


# =============================================================================
# 6. AFFICHAGE DES RÉSULTATS
# =============================================================================

def afficher_resultats(solution: Dict, Z: float, label: str = "HEURISTIQUE") -> None:
    """Affiche le planning complet d'une solution."""
    sel    = solution["selection"]
    aff    = solution["quais"]
    _, details = evaluer(solution)
    ordres = solution.get("_ordres", {})
    exclus = [c for c in CAMIONS if c not in sel]

    print("\n" + "=" * 95)
    print(f"  RÉSULTATS — {label}")
    print("=" * 95)
    print(f"  Valeur objective Z = {Z:.2f}")
    print()
    print(f"  ✔  Sélectionnés ({len(sel)}) : {', '.join(sorted(sel))}")
    print(f"  ✘  Exclus       ({len(exclus)})  : {', '.join(sorted(exclus))}")

    W = {
        "camion"  :  6, "trans"   : 13, "eco"     :  4,
        "arrivee" :  7, "debut"   :  7, "attente" :  7,
        "fin"     :  7, "limite"  :  7, "retard"  :  7, "penalite":  9,
    }

    def row_str(cam, trans, eco, arr, deb, att, fin, lim, ret, pen):
        return (
            f"  │  {str(cam):<{W['camion']}}  {str(trans):<{W['trans']}}  "
            f"{str(eco):>{W['eco']}}  {str(arr):>{W['arrivee']}}  "
            f"{str(deb):>{W['debut']}}  {str(att):>{W['attente']}}  "
            f"{str(fin):>{W['fin']}}  {str(lim):>{W['limite']}}  "
            f"{str(ret):>{W['retard']}}  {str(pen):>{W['penalite']}}"
        )

    header = row_str("Camion", "Transporteur", "Éco",
                     "Arrivée", "Début", "Attente", "Fin", "Limite", "Retard", "Pénalité")
    sep    = row_str(*["─" * w for w in W.values()])
    line_w = len(sep) - 4

    print()
    for q in range(1, Q + 1):
        ordre_q = ordres.get(q, sorted(
            [c for c in sel if aff[c] == q],
            key=lambda c: details.get(c, {}).get("d", 0)
        ))

        # Recalcul des débuts dans l'ordre optimal retenu (avec continuité du quai)
        tl = 0
        details_quai = {}
        for c in ordre_q:
            debut = max(CAMIONS[c]["Ta"], tl)
            fin   = debut + CAMIONS[c]["Td"]
            details_quai[c] = {
                "d":      debut,
                "r":      max(0.0, fin - CAMIONS[c]["Tl"]),
                "attente": debut - CAMIONS[c]["Ta"],
                "fin":    fin,
            }
            tl = fin

        titre = f" QUAI {q} ({len(ordre_q)} camion{'s' if len(ordre_q) != 1 else ''}) "
        print(f"  ┌─{titre}{'─' * (line_w - len(titre) - 1)}")
        print(header)
        print(sep)

        for c in ordre_q:
            d   = details_quai.get(c, details.get(c, {}))
            Ta  = CAMIONS[c]["Ta"]
            di  = d.get("d",      Ta)
            ri  = d.get("r",      0)
            fin = d.get("fin",    di + CAMIONS[c]["Td"])
            at  = d.get("attente", 0)
            Pi  = CAMIONS[c]["P"]
            Tl  = CAMIONS[c]["Tl"]

            print(row_str(
                c,
                CAMIONS[c]["transporteur"],
                CAMIONS[c]["S"],
                min_to_hhmm(Ta),
                min_to_hhmm(di),
                f"{int(round(at))}m",
                min_to_hhmm(fin),
                min_to_hhmm(Tl),
                f"{int(round(ri))}m",
                f"{int(round(Pi * ri))} ₮",
            ))

        print(f"  └{'─' * (line_w + 2)}")
        print()

    score_eco   = sum(CAMIONS[c]["S"]                          for c in sel)
    penalites   = sum(CAMIONS[c]["P"] * details.get(c, {}).get("r", 0) for c in sel)
    attente_tot = sum(details.get(c, {}).get("attente", 0)     for c in sel)

    print(f"  Décomposition de Z :")
    print(f"    + Score éco total     : {score_eco:.0f}")
    print(f"    − Pénalités retard    : {penalites:.0f} TND")
    print(f"    − Temps attente total : {attente_tot:.0f} min")
    print(f"    {'═' * 34}")
    print(f"      Z                   = {Z:.2f}")
    print()


def afficher_convergence(historique: List[float]) -> None:
    """Affiche une courbe de convergence ASCII de l'AG."""
    print("\n" + "─" * 65)
    print("  COURBE DE CONVERGENCE DE L'ALGORITHME GÉNÉTIQUE")
    print("─" * 65)

    Z_min    = min(historique)
    Z_max    = max(historique)
    hauteur  = 10
    largeur  = min(len(historique), 60)
    pas      = max(1, len(historique) // largeur)
    ech      = [historique[i] for i in range(0, len(historique), pas)][:largeur]

    def norm(v):
        if Z_max == Z_min:
            return hauteur // 2
        return int((v - Z_min) / (Z_max - Z_min) * (hauteur - 1))

    grille = [[" "] * len(ech) for _ in range(hauteur)]
    for col, v in enumerate(ech):
        grille[hauteur - 1 - norm(v)][col] = "█"

    for i, ligne in enumerate(grille):
        val = Z_max - (Z_max - Z_min) * i / (hauteur - 1)
        print(f"  {val:9.1f} │ {''.join(ligne)}")

    print(f"  {'':>9} └{'─' * len(ech)}")
    print(f"  {'':>11}Gén. 0{' ' * (len(ech) - 18)}Gén. {len(historique) - 1}")
    print()


# =============================================================================
# 7. POINT D'ENTRÉE
# =============================================================================

def main():
    print("=" * 65)
    print("  BALLET DES TRANSPORTEURS GWD — MÉTHODES HEURISTIQUES")
    print("  Greedy + Algorithme Génétique + Optimisation Ordre Quais")
    print("=" * 65)

    # 1. Chargement des données depuis l'Excel
    charger_depuis_excel()
    print(f"\n[INFO] N={N}, K={K}, Q={Q}\n")

    t_total = time.time()

    # 2. Phase 1 : Greedy
    sol_greedy = greedy()
    Z_greedy, _ = evaluer(sol_greedy)
    afficher_resultats(sol_greedy, Z_greedy, "SOLUTION GREEDY (initiale)")

    # 3. Phase 2 : Algorithme Génétique
    sol_ag, Z_ag, historique = algorithme_genetique(
        solution_initiale = sol_greedy,
        taille_pop        = 80,
        n_generations     = 300,
        taux_croisement   = 0.85,
        taux_mutation     = 0.30,
        taille_tournoi    = 3,
        elitisme          = 3,
        patience          = 80,
        graine            = 42,
    )

    afficher_resultats(sol_ag, Z_ag, "SOLUTION ALGORITHME GÉNÉTIQUE (finale)")
    afficher_convergence(historique)

    # 4. Résumé comparatif
    print("─" * 65)
    print("  RÉSUMÉ COMPARATIF")
    print("─" * 65)
    amelioration = Z_ag - Z_greedy
    pct          = (amelioration / abs(Z_greedy) * 100) if Z_greedy != 0 else 0
    print(f"  Z  Greedy               : {Z_greedy:>10.2f}")
    print(f"  Z  Algorithme Génétique : {Z_ag:>10.2f}  (Δ = {amelioration:+.2f}"
          f" / {pct:+.1f}% vs Greedy)")
    print(f"  Temps total             : {time.time() - t_total:.3f}s")
    print("─" * 65)


if __name__ == "__main__":
    main()